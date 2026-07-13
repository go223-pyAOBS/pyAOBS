from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Any
from collections import Counter


TARGET_COLUMNS = [
    "rock_type",
    "abbrev",
    "岩石属性",
    "density",
    "det_den",
    "vp",
    "dvp",
    "vs",
    "dvs",
    "pressure",
    "pressure_unit",
    "temperature",
    "felsic_or_mafic",
    "rock_facies",
    "sio2_wt",
    "source",
    "method",
]


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _to_float(value: Any) -> float | None:
    s = _clean_text(value)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _normalize_density(value: Any, unit: str) -> float | None:
    x = _to_float(value)
    if x is None:
        return None
    u = _clean_text(unit).lower().replace(" ", "")
    if u in {"kg/m3", "kgm3", "kg_m3"}:
        return x
    if u in {"g/cm3", "gcm3", "g_cm3"}:
        return x * 1000.0
    # No unit provided: infer by magnitude
    return x * 1000.0 if x < 20 else x


def _normalize_pressure(value: Any, unit: str) -> float | None:
    x = _to_float(value)
    if x is None:
        return None
    u = _clean_text(unit).lower().replace(" ", "")
    if u in {"mpa"}:
        return x
    if u in {"gpa"}:
        return x * 1000.0
    # No unit provided: infer by magnitude
    # Typical lab pressure in this workflow is around hundreds of MPa.
    # Values <= 5 are likely GPa.
    return x * 1000.0 if x <= 5 else x


def _pick_first_text(raw: dict[str, str], aliases: list[str]) -> str:
    lower_map = {str(k).strip().lower(): k for k in raw.keys()}
    for alias in aliases:
        key = lower_map.get(alias.lower())
        if key is None:
            continue
        val = _clean_text(raw.get(key, ""))
        if val:
            return val
    return ""


def _pick_first_float(raw: dict[str, str], aliases: list[str]) -> float | None:
    lower_map = {str(k).strip().lower(): k for k in raw.keys()}
    for alias in aliases:
        key = lower_map.get(alias.lower())
        if key is None:
            continue
        v = _to_float(raw.get(key, ""))
        if v is not None:
            return v
    return None


def _normalize_base_row(raw: dict[str, str], base_name: str) -> dict[str, str] | None:
    rock_type = _pick_first_text(raw, ["rock_type", "rock type", "type", "岩石类型", "岩性"])
    if not rock_type:
        return None

    vp = _pick_first_float(raw, ["vp", "v_p", "p-wave", "pwave", "v"])
    if vp is None:
        return None
    vs = _pick_first_float(raw, ["vs", "v_s", "s-wave", "swave"])

    density_raw = _pick_first_float(raw, ["density", "rho", "密度"])
    density = None
    if density_raw is not None:
        density = _normalize_density(density_raw, "")

    pressure_raw = _pick_first_float(raw, ["pressure", "p", "压力"])
    pressure = _normalize_pressure(pressure_raw, "MPa") if pressure_raw is not None else 200.0

    out = {col: "" for col in TARGET_COLUMNS}
    out["rock_type"] = rock_type
    out["abbrev"] = _pick_first_text(raw, ["abbrev", "abbre", "abbr"])
    out["岩石属性"] = _pick_first_text(raw, ["岩石属性", "中文", "中文名"])
    if density is not None:
        out["density"] = f"{round(density):.0f}"
    out["vp"] = f"{vp:.3f}"
    out["vs"] = f"{vs:.3f}" if vs is not None else ""
    out["pressure"] = f"{pressure:.0f}"
    out["pressure_unit"] = "MPa"
    out["temperature"] = _pick_first_text(raw, ["temperature", "temp", "温度"]) or "25"
    out["felsic_or_mafic"] = _pick_first_text(raw, ["felsic_or_mafic", "felsic or mafic", "composition"])
    out["rock_facies"] = _pick_first_text(raw, ["rock_facies", "rock facies", "facies"])
    sio2 = _pick_first_float(raw, ["sio2_wt", "sio2", "sio2, wt.%", "sio2 wt"])
    out["sio2_wt"] = f"{sio2:.3f}" if sio2 is not None else ""
    out["source"] = _pick_first_text(raw, ["source", "reference", "citation"]) or base_name
    out["method"] = _pick_first_text(raw, ["method", "measurement_method"]) or "legacy_import"
    return out


def _canonical_key(row: dict[str, str]) -> tuple[str, ...]:
    def rf(name: str, ndigits: int) -> str:
        v = _to_float(row.get(name, ""))
        if v is None:
            return ""
        return f"{round(v, ndigits):.{ndigits}f}"

    return (
        _clean_text(row.get("rock_type", "")).lower(),
        rf("density", 1),
        rf("vp", 3),
        rf("vs", 3),
        rf("pressure", 1),
        rf("temperature", 1),
        _clean_text(row.get("source", "")).lower(),
        _clean_text(row.get("method", "")).lower(),
    )


def _normalize_incoming_row(raw: dict[str, str]) -> dict[str, str] | None:
    rock_type = _clean_text(raw.get("rock_type", ""))
    vp = _to_float(raw.get("vp", ""))
    vs = _to_float(raw.get("vs", ""))
    pressure = _normalize_pressure(
        raw.get("pressure", ""),
        raw.get("pressure_unit", ""),
    )
    temperature = _to_float(raw.get("temperature", ""))
    density = _normalize_density(raw.get("density", ""), raw.get("density_unit", ""))
    source = _clean_text(raw.get("source", ""))
    method = _clean_text(raw.get("method", ""))

    if (
        not rock_type
        or vp is None
        or density is None
        or pressure is None
        or temperature is None
        or not source
        or not method
    ):
        return None

    out: dict[str, str] = {}
    for col in TARGET_COLUMNS:
        out[col] = ""

    out["rock_type"] = rock_type
    out["abbrev"] = _clean_text(raw.get("abbrev", ""))
    out["岩石属性"] = _clean_text(raw.get("岩石属性", ""))
    out["density"] = f"{round(density):.0f}"
    out["det_den"] = _clean_text(raw.get("det_den", ""))
    out["vp"] = f"{vp:.3f}"
    out["dvp"] = _clean_text(raw.get("dvp", ""))
    out["vs"] = f"{vs:.3f}" if vs is not None else ""
    out["dvs"] = _clean_text(raw.get("dvs", ""))
    out["pressure"] = f"{pressure:.0f}"
    out["pressure_unit"] = "MPa"
    out["temperature"] = f"{temperature:.0f}"
    out["felsic_or_mafic"] = _clean_text(raw.get("felsic_or_mafic", ""))
    out["rock_facies"] = _clean_text(raw.get("rock_facies", ""))
    sio2 = _to_float(raw.get("sio2_wt", ""))
    out["sio2_wt"] = f"{sio2:.3f}" if sio2 is not None else ""
    out["source"] = source
    out["method"] = method
    return out


def _missing_required_fields(raw: dict[str, str]) -> list[str]:
    missing: list[str] = []
    rock_type = _clean_text(raw.get("rock_type", ""))
    vp = _to_float(raw.get("vp", ""))
    pressure = _normalize_pressure(raw.get("pressure", ""), raw.get("pressure_unit", ""))
    temperature = _to_float(raw.get("temperature", ""))
    density = _normalize_density(raw.get("density", ""), raw.get("density_unit", ""))
    source = _clean_text(raw.get("source", ""))
    method = _clean_text(raw.get("method", ""))

    if not rock_type:
        missing.append("rock_type")
    if vp is None:
        missing.append("vp")
    if density is None:
        missing.append("density")
    if pressure is None:
        missing.append("pressure")
    if temperature is None:
        missing.append("temperature")
    if not source:
        missing.append("source")
    if not method:
        missing.append("method")
    return missing


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _read_excel(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "Reading .xlsx base file requires openpyxl. Please install: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        values = ws.values
        header_row = next(values, None)
        if header_row is None:
            return []
        headers = [_clean_text(v) for v in header_row]
        rows: list[dict[str, str]] = []
        for r in values:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _clean_text(r[i] if i < len(r) else "")
            rows.append(row)
        return rows
    finally:
        wb.close()


def _read_table(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel(path)
    return _read_csv(path)


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TARGET_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in TARGET_COLUMNS})


def merge(base_path: Path, incoming_path: Path, out_path: Path) -> tuple[int, int, dict[str, int]]:
    if base_path.exists():
        raw_base_rows = _read_table(base_path)
        base_rows = []
        for raw in raw_base_rows:
            normalized = _normalize_base_row(raw, base_path.name)
            if normalized is not None:
                base_rows.append(normalized)
    else:
        base_rows = []
    normalized_new: list[dict[str, str]] = []
    skipped = 0
    missing_counter: Counter[str] = Counter()
    for raw in _read_table(incoming_path):
        row = _normalize_incoming_row(raw)
        if row is None:
            skipped += 1
            missing_counter.update(_missing_required_fields(raw))
            continue
        normalized_new.append(row)

    existing_keys = {_canonical_key(r) for r in base_rows}
    added = 0
    for row in normalized_new:
        key = _canonical_key(row)
        if key in existing_keys:
            continue
        base_rows.append(row)
        existing_keys.add(key)
        added += 1

    _write_csv(out_path, base_rows)
    return added, skipped, dict(missing_counter)


def _default_base_database() -> Path:
    """返回默认基库路径，优先使用 utils/rockdata。"""
    utils_dir = Path(__file__).parent
    candidates = [
        utils_dir / "rockdata" / "rocks_merged.csv",
        utils_dir / "rockdata" / "rocks.csv",
        utils_dir / "rocks.csv",
    ]
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


def _fallback_output_path(target: Path) -> Path:
    """当默认输出文件不可写时，生成同目录兜底文件名。"""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return target.with_name(f"{target.stem}_{stamp}{target.suffix}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge external rock property CSV into utils/rockdata database."
    )
    parser.add_argument(
        "--incoming",
        required=True,
        help="Path to external CSV (recommended: external_rock_data_template.csv format).",
    )
    parser.add_argument(
        "--base",
        default="",
        help=(
            "Base database path. Default auto-selects "
            "utils/rockdata/rocks_merged.csv -> utils/rockdata/rocks.csv -> utils/rocks.csv."
        ),
    )
    parser.add_argument(
        "--out",
        default="",
        help="Output path. Default overwrites --base.",
    )
    args = parser.parse_args()

    base = (
        Path(args.base).expanduser().resolve()
        if args.base
        else _default_base_database().resolve()
    )
    incoming = Path(args.incoming).expanduser().resolve()
    out = Path(args.out).expanduser().resolve() if args.out else base

    if args.base and not base.exists():
        raise FileNotFoundError(f"Base file not found: {base}")
    if not incoming.exists():
        raise FileNotFoundError(f"Incoming file not found: {incoming}")
    out.parent.mkdir(parents=True, exist_ok=True)
    if (not args.base) and (not base.exists()):
        print(f"base_init=new_file:{base}")

    try:
        added, skipped, missing_stats = merge(base, incoming, out)
    except PermissionError:
        if args.out:
            raise
        alt_out = _fallback_output_path(out)
        print(f"write_fallback_from={out}")
        print(f"write_fallback_to={alt_out}")
        added, skipped, missing_stats = merge(base, incoming, alt_out)
        out = alt_out

    print(f"merged_to={out}")
    print(f"added={added}")
    print(f"skipped_invalid={skipped}")
    if skipped:
        print(f"skipped_missing_stats={missing_stats}")


if __name__ == "__main__":
    main()
