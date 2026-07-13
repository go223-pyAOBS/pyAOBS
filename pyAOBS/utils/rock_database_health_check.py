from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path
from typing import Any


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _to_float(value: Any) -> float | None:
    s = _clean_text(value).replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _read_excel(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Please install: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        values = ws.values
        header_row = next(values, None)
        if header_row is None:
            return []
        headers = [_clean_text(v) for v in header_row]
        rows: list[dict[str, str]] = []
        for r in values:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _clean_text(r[i] if i < len(r) else "")
            rows.append(row)
        return rows
    finally:
        wb.close()


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_excel(path)
    return _read_csv(path)


def _col(row: dict[str, str], *aliases: str) -> str:
    lower_map = {str(k).strip().lower(): k for k in row.keys()}
    for alias in aliases:
        key = lower_map.get(alias.lower())
        if key is None:
            continue
        val = _clean_text(row.get(key, ""))
        if val:
            return val
    return ""


def main() -> None:
    parser = argparse.ArgumentParser(description="Health check rock database table.")
    parser.add_argument("--db", required=True, help="Path to rock database CSV/XLSX.")
    parser.add_argument("--top", type=int, default=15, help="Top N categories to print.")
    args = parser.parse_args()

    db_path = Path(args.db).expanduser().resolve()
    if not db_path.exists():
        raise FileNotFoundError(f"Database file not found: {db_path}")

    rows = _read_table(db_path)
    n = len(rows)
    print(f"db={db_path}")
    print(f"rows={n}")
    if n == 0:
        return

    rock_types: Counter[str] = Counter()
    sources: Counter[str] = Counter()
    pressures: Counter[str] = Counter()
    facies: Counter[str] = Counter()
    felsic_mafic: Counter[str] = Counter()

    miss_vp = miss_vs = miss_den = miss_press = miss_source = 0
    have_sio2 = 0
    sio2_values: list[float] = []

    for row in rows:
        rock = _col(row, "rock_type", "rock type", "type", "岩性")
        if rock:
            rock_types[rock] += 1

        vp = _to_float(_col(row, "vp", "v_p", "p-wave", "v"))
        vs = _to_float(_col(row, "vs", "v_s", "s-wave"))
        den = _to_float(_col(row, "density", "rho", "密度"))
        press = _to_float(_col(row, "pressure", "p", "压力"))
        src = _col(row, "source", "reference", "citation")
        fac = _col(row, "rock_facies", "rock facies", "facies")
        comp = _col(row, "felsic_or_mafic", "felsic or mafic")
        sio2 = _to_float(_col(row, "sio2_wt", "sio2", "sio2, wt.%"))

        if vp is None:
            miss_vp += 1
        if vs is None:
            miss_vs += 1
        if den is None:
            miss_den += 1
        if press is None:
            miss_press += 1
        if not src:
            miss_source += 1

        if src:
            sources[src] += 1
        if press is not None:
            pressures[f"{press:.3f}"] += 1
        if fac:
            facies[fac] += 1
        if comp:
            felsic_mafic[comp] += 1
        if sio2 is not None:
            have_sio2 += 1
            sio2_values.append(sio2)

    def pct(x: int) -> str:
        return f"{(x / n * 100.0):.1f}%"

    print(f"missing_vp={miss_vp} ({pct(miss_vp)})")
    print(f"missing_vs={miss_vs} ({pct(miss_vs)})")
    print(f"missing_density={miss_den} ({pct(miss_den)})")
    print(f"missing_pressure={miss_press} ({pct(miss_press)})")
    print(f"missing_source={miss_source} ({pct(miss_source)})")
    print(f"rock_type_unique={len(rock_types)}")
    print(f"source_unique={len(sources)}")
    print(f"pressure_unique={len(pressures)}")
    print(f"felsic_or_mafic_unique={len(felsic_mafic)}")
    print(f"rock_facies_unique={len(facies)}")
    print(f"sio2_present={have_sio2} ({pct(have_sio2)})")
    if sio2_values:
        print(
            f"sio2_range=[{min(sio2_values):.3f}, {max(sio2_values):.3f}] "
            f"median={sorted(sio2_values)[len(sio2_values)//2]:.3f}"
        )

    top_n = max(1, args.top)
    print(f"top_rock_types={rock_types.most_common(top_n)}")
    print(f"top_sources={sources.most_common(top_n)}")
    print(f"top_pressures={pressures.most_common(top_n)}")
    print(f"top_felsic_or_mafic={felsic_mafic.most_common(top_n)}")
    print(f"top_rock_facies={facies.most_common(top_n)}")


if __name__ == "__main__":
    main()
