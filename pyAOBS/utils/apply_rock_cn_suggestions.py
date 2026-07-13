from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any
import zipfile
import xml.etree.ElementTree as ET


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _norm_key(value: str) -> str:
    s = _clean_text(value).lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^a-z0-9\u4e00-\u9fff\-/,() ]+", "", s)
    return s.strip()


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return _xlsx_rows_fallback(path)

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        values = ws.values
        header_row = next(values, None)
        if header_row is None:
            return []
        headers = [_clean_text(v) for v in header_row]
        rows: list[dict[str, str]] = []
        for r in values:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _clean_text(r[i] if i < len(r) else "")
            rows.append(row)
        return rows
    finally:
        wb.close()


def _xlsx_rows_fallback(path: Path) -> list[dict[str, str]]:
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    relns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with zipfile.ZipFile(path) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        first_sheet = wb.findall(f".//{ns}sheet")[0]
        rel_id = first_sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", ""
        )
        rel_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            r.attrib.get("Id", ""): r.attrib.get("Target", "")
            for r in rel_xml.findall(f".//{relns}Relationship")
        }
        target = rel_map.get(rel_id, "").lstrip("/")
        if not target:
            return []
        sheet_xml_path = f"xl/{target}" if not target.startswith("xl/") else target
        sheet = ET.fromstring(z.read(sheet_xml_path))

        shared_strings: list[str] = []
        try:
            sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in sst.findall(f".//{ns}si"):
                txt = "".join(t.text or "" for t in si.findall(f".//{ns}t"))
                shared_strings.append(txt)
        except KeyError:
            pass

        def col_to_idx(col: str) -> int:
            n = 0
            for ch in col:
                if "A" <= ch <= "Z":
                    n = n * 26 + (ord(ch) - 64)
            return n - 1

        all_rows: list[list[str]] = []
        for row in sheet.findall(f".//{ns}sheetData/{ns}row"):
            cells = row.findall(f"{ns}c")
            values: dict[int, str] = {}
            max_idx = -1
            for c in cells:
                ref = c.attrib.get("r", "A1")
                m = re.match(r"([A-Z]+)", ref)
                if not m:
                    continue
                idx = col_to_idx(m.group(1))
                if idx > max_idx:
                    max_idx = idx
                v = c.find(f"{ns}v")
                if v is None:
                    values[idx] = ""
                    continue
                raw = v.text or ""
                if c.attrib.get("t") == "s" and raw.isdigit():
                    i = int(raw)
                    txt = shared_strings[i] if 0 <= i < len(shared_strings) else raw
                else:
                    txt = raw
                values[idx] = _clean_text(txt)
            if max_idx < 0:
                continue
            line = [values.get(i, "") for i in range(max_idx + 1)]
            all_rows.append(line)

        if not all_rows:
            return []
        headers = [_clean_text(v) for v in all_rows[0]]
        out: list[dict[str, str]] = []
        for r in all_rows[1:]:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _clean_text(r[i] if i < len(r) else "")
            out.append(row)
        return out


def _read_suggestions(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _xlsx_rows(path)
    return _read_csv(path)


def _is_true(value: str) -> bool:
    return _clean_text(value).lower() in {"y", "yes", "true", "1", "ok"}


def apply_suggestions(
    db_csv: Path,
    suggestions_path: Path,
    output_csv: Path,
    backup: bool = True,
) -> tuple[int, int, int]:
    rows = _read_csv(db_csv)
    if not rows:
        _write_csv(output_csv, [], [])
        return 0, 0, 0
    original_rows = [dict(r) for r in rows]

    sugg_rows = _read_suggestions(suggestions_path)
    mapping: dict[str, str] = {}
    for s in sugg_rows:
        rock_type = _clean_text(s.get("rock_type", ""))
        accepted = _is_true(_clean_text(s.get("accepted", "")))
        final_cn = _clean_text(s.get("final_cn", ""))
        suggested_cn = _clean_text(s.get("suggested_cn", ""))
        if not rock_type:
            continue
        chosen = final_cn or suggested_cn
        if accepted and chosen:
            mapping[_norm_key(rock_type)] = chosen
        elif final_cn and not _clean_text(s.get("accepted", "")):
            # If user only filled final_cn but left accepted empty, still accept it.
            mapping[_norm_key(rock_type)] = final_cn

    if "岩石属性" not in rows[0]:
        for r in rows:
            r["岩石属性"] = ""
    fieldnames = list(rows[0].keys())

    applied = 0
    touched_keys: set[str] = set()
    for r in rows:
        rock = _clean_text(r.get("rock_type", ""))
        key = _norm_key(rock)
        if not key:
            continue
        cn = mapping.get(key)
        if not cn:
            continue
        if _clean_text(r.get("岩石属性", "")) != cn:
            r["岩石属性"] = cn
            applied += 1
        touched_keys.add(key)

    if backup and db_csv.resolve() == output_csv.resolve():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = db_csv.with_name(f"{db_csv.stem}.backup_{stamp}{db_csv.suffix}")
        _write_csv(backup_path, original_rows, fieldnames)

    _write_csv(output_csv, rows, fieldnames)
    return len(mapping), len(touched_keys), applied


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Apply reviewed Chinese-name suggestions back to rocks_merged.csv."
    )
    parser.add_argument("--db", required=True, help="Path to rocks_merged.csv")
    parser.add_argument(
        "--suggestions",
        required=True,
        help="Path to rock_type_cn_suggestions.csv/xlsx",
    )
    parser.add_argument(
        "--out",
        default="",
        help="Output CSV path. Default overwrites --db.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create backup when overwriting --db.",
    )
    args = parser.parse_args()

    db = Path(args.db).expanduser().resolve()
    suggestions = Path(args.suggestions).expanduser().resolve()
    out = Path(args.out).expanduser().resolve() if args.out else db

    if not db.exists():
        raise FileNotFoundError(f"Database file not found: {db}")
    if not suggestions.exists():
        raise FileNotFoundError(f"Suggestions file not found: {suggestions}")

    mapping_size, matched_types, applied_rows = apply_suggestions(
        db_csv=db,
        suggestions_path=suggestions,
        output_csv=out,
        backup=not args.no_backup,
    )
    print(f"output={out}")
    print(f"accepted_mapping_count={mapping_size}")
    print(f"matched_rock_type_count={matched_types}")
    print(f"applied_row_count={applied_rows}")


if __name__ == "__main__":
    main()
