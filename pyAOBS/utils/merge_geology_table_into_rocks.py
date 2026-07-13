from __future__ import annotations

import argparse
import csv
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _norm_key(value: str) -> str:
    s = _clean_text(value).lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^a-z0-9\u4e00-\u9fff\-/,() ]+", "", s)
    return s.strip()


def _norm_key_loose(value: str) -> str:
    s = _norm_key(value)
    s = s.replace("-", " ").replace("/", " ").replace(",", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _read_csv(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    raise ValueError(f"Failed to read CSV: {path}")


def _write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return _read_xlsx_fallback(path)

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        values = ws.values
        header_row = next(values, None)
        if header_row is None:
            return []
        headers = [_clean_text(v) for v in header_row]
        rows: list[dict[str, str]] = []
        for r in values:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _clean_text(r[i] if i < len(r) else "")
            rows.append(row)
        return rows
    finally:
        wb.close()


def _read_xlsx_fallback(path: Path) -> list[dict[str, str]]:
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    relns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with zipfile.ZipFile(path) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        first_sheet = wb.findall(f".//{ns}sheet")[0]
        rel_id = first_sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", ""
        )
        rel_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            r.attrib.get("Id", ""): r.attrib.get("Target", "")
            for r in rel_xml.findall(f".//{relns}Relationship")
        }
        target = rel_map.get(rel_id, "").lstrip("/")
        if not target:
            return []
        sheet_xml_path = f"xl/{target}" if not target.startswith("xl/") else target
        sheet = ET.fromstring(z.read(sheet_xml_path))

        shared_strings: list[str] = []
        try:
            sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in sst.findall(f".//{ns}si"):
                txt = "".join(t.text or "" for t in si.findall(f".//{ns}t"))
                shared_strings.append(txt)
        except KeyError:
            pass

        def col_to_idx(col: str) -> int:
            n = 0
            for ch in col:
                if "A" <= ch <= "Z":
                    n = n * 26 + (ord(ch) - 64)
            return n - 1

        all_rows: list[list[str]] = []
        for row in sheet.findall(f".//{ns}sheetData/{ns}row"):
            values: dict[int, str] = {}
            max_idx = -1
            for c in row.findall(f"{ns}c"):
                ref = c.attrib.get("r", "A1")
                m = re.match(r"([A-Z]+)", ref)
                if not m:
                    continue
                idx = col_to_idx(m.group(1))
                if idx > max_idx:
                    max_idx = idx
                v = c.find(f"{ns}v")
                if v is None:
                    values[idx] = ""
                    continue
                raw = v.text or ""
                if c.attrib.get("t") == "s" and raw.isdigit():
                    i = int(raw)
                    txt = shared_strings[i] if 0 <= i < len(shared_strings) else raw
                else:
                    txt = raw
                values[idx] = _clean_text(txt)
            if max_idx >= 0:
                all_rows.append([values.get(i, "") for i in range(max_idx + 1)])

        if not all_rows:
            return []
        headers = [_clean_text(v) for v in all_rows[0]]
        rows: list[dict[str, str]] = []
        for r in all_rows[1:]:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _clean_text(r[i] if i < len(r) else "")
            rows.append(row)
        return rows


def _read_table(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return _read_csv(path)


def _find_col(row: dict[str, str], aliases: list[str]) -> str:
    key_map = {str(k).strip().lower(): k for k in row.keys()}
    for alias in aliases:
        k = key_map.get(alias.lower())
        if k:
            return k
    # relaxed match by contained text
    for key in row.keys():
        lk = str(key).strip().lower()
        if any(alias.lower() in lk for alias in aliases):
            return key
    return ""


def _build_geology_mapping(rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    strict: dict[str, dict[str, str]] = {}
    loose: dict[str, dict[str, str]] = {}
    if not rows:
        return strict, loose

    sample = rows[0]
    rock_col = _find_col(sample, ["rock_type", "rock type", "岩石类型"])
    cn_col = _find_col(sample, ["岩石属性", "中文名称", "中文"])
    class_col = _find_col(sample, ["岩石分类"])
    grade_col = _find_col(sample, ["变质程度"])
    meaning_col = _find_col(sample, ["地质意义"])
    foma_col = _find_col(sample, ["岩性类别", "felsic_or_mafic", "felsic or mafic"])

    for r in rows:
        rock = _clean_text(r.get(rock_col, ""))
        if not rock:
            continue
        payload = {
            "岩石属性": _clean_text(r.get(cn_col, "")) if cn_col else "",
            "岩石分类": _clean_text(r.get(class_col, "")) if class_col else "",
            "变质程度": _clean_text(r.get(grade_col, "")) if grade_col else "",
            "地质意义": _clean_text(r.get(meaning_col, "")) if meaning_col else "",
            "felsic_or_mafic": _clean_text(r.get(foma_col, "")) if foma_col else "",
        }
        sk = _norm_key(rock)
        lk = _norm_key_loose(rock)
        if sk and sk not in strict:
            strict[sk] = payload
        if lk and lk not in loose:
            loose[lk] = payload
    return strict, loose


def merge_geology_table(
    rocks_csv: Path,
    geology_table: Path,
    output_csv: Path,
    overwrite_existing: bool = True,
    backup: bool = True,
) -> tuple[int, int]:
    rocks_rows = _read_csv(rocks_csv)
    if not rocks_rows:
        _write_csv(output_csv, [], [])
        return 0, 0

    geology_rows = _read_table(geology_table)
    strict, loose = _build_geology_mapping(geology_rows)

    fieldnames = list(rocks_rows[0].keys())
    for col in ["岩石属性", "岩石分类", "变质程度", "地质意义", "felsic_or_mafic"]:
        if col not in fieldnames:
            fieldnames.append(col)

    updated_rows = 0
    matched_rock_types: set[str] = set()
    for row in rocks_rows:
        rock = _clean_text(row.get("rock_type", ""))
        if not rock:
            continue
        key = _norm_key(rock)
        payload = strict.get(key) or loose.get(_norm_key_loose(rock))
        if not payload:
            continue
        matched_rock_types.add(key)
        changed = False
        for col, val in payload.items():
            if not val:
                continue
            current = _clean_text(row.get(col, ""))
            if (not current) or overwrite_existing:
                if current != val:
                    row[col] = val
                    changed = True
        if changed:
            updated_rows += 1

    if backup and rocks_csv.resolve() == output_csv.resolve():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = rocks_csv.with_name(f"{rocks_csv.stem}.backup_before_geology_merge_{stamp}{rocks_csv.suffix}")
        _write_csv(backup_path, rocks_rows, fieldnames)

    _write_csv(output_csv, rocks_rows, fieldnames)
    return len(matched_rock_types), updated_rows


def _fallback_output_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_geology_merged_{stamp}{path.suffix}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge geology meaning table into rocks_merged.csv."
    )
    parser.add_argument("--rocks", required=True, help="Path to rocks_merged.csv")
    parser.add_argument("--table", required=True, help="Path to geology table (csv/xlsx)")
    parser.add_argument("--out", default="", help="Output path. Default overwrite --rocks")
    parser.add_argument(
        "--no-overwrite-existing",
        action="store_true",
        help="Only fill empty fields; do not overwrite existing non-empty values.",
    )
    parser.add_argument("--no-backup", action="store_true", help="Disable backup when overwrite --rocks")
    args = parser.parse_args()

    rocks = Path(args.rocks).expanduser().resolve()
    table = Path(args.table).expanduser().resolve()
    out = Path(args.out).expanduser().resolve() if args.out else rocks
    if not rocks.exists():
        raise FileNotFoundError(f"Rocks file not found: {rocks}")
    if not table.exists():
        raise FileNotFoundError(f"Geology table not found: {table}")

    try:
        matched, updated = merge_geology_table(
            rocks_csv=rocks,
            geology_table=table,
            output_csv=out,
            overwrite_existing=not args.no_overwrite_existing,
            backup=not args.no_backup,
        )
    except PermissionError:
        if args.out:
            raise
        alt_out = _fallback_output_path(out)
        print(f"write_fallback_from={out}")
        print(f"write_fallback_to={alt_out}")
        matched, updated = merge_geology_table(
            rocks_csv=rocks,
            geology_table=table,
            output_csv=alt_out,
            overwrite_existing=not args.no_overwrite_existing,
            backup=False,
        )
        out = alt_out
    print(f"output={out}")
    print(f"matched_rock_type_count={matched}")
    print(f"updated_row_count={updated}")


if __name__ == "__main__":
    main()
