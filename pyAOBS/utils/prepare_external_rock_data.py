from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any


TEMPLATE_COLUMNS = [
    "rock_type",
    "abbrev",
    "岩石属性",
    "density",
    "density_unit",
    "det_den",
    "vp",
    "dvp",
    "vs",
    "dvs",
    "pressure",
    "pressure_unit",
    "temperature",
    "felsic_or_mafic",
    "rock_facies",
    "sio2_wt",
    "source",
    "method",
]


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _norm_header(name: str) -> str:
    s = _clean_text(name).lower()
    s = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _to_float(value: Any) -> float | None:
    s = _clean_text(value)
    if not s:
        return None
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-+eE]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    last_error: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                headers = list(reader.fieldnames or [])
                rows = list(reader)
            return headers, rows
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    raise ValueError(f"Failed to read CSV: {path}")


def _read_excel(path: Path, sheet_name: str = "") -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Please install: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                names = ", ".join(wb.sheetnames)
                raise ValueError(
                    f"Sheet '{sheet_name}' not found in {path.name}. Available: {names}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        all_rows = [list(r) if r is not None else [] for r in ws.values]
        if not all_rows:
            return [], []

        def row_score(row: list[Any]) -> tuple[int, int]:
            texts = [_clean_text(v) for v in row]
            non_empty = [t for t in texts if t]
            if not non_empty:
                return (0, 0)
            joined = " ".join(non_empty).lower()
            keywords = [
                "rock type",
                "rock_type",
                "density",
                "vp",
                "vs",
                "pressure",
                "temperature",
                "sample",
            ]
            hit = sum(1 for kw in keywords if kw in joined)
            return (hit, len(non_empty))

        scan_rows = all_rows[:80]
        best_idx = 0
        best = (-1, -1)
        for idx, row in enumerate(scan_rows):
            score = row_score(row)
            if score > best:
                best = score
                best_idx = idx

        headers = [_clean_text(v) for v in all_rows[best_idx]]
        rows: list[dict[str, str]] = []
        for r in all_rows[best_idx + 1 :]:
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                value = r[i] if i < len(r) else ""
                row[h] = _clean_text(value)
            rows.append(row)
        return headers, rows
    finally:
        wb.close()


def _read_excel_all_rows(path: Path, sheet_name: str = "") -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Please install: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                names = ", ".join(wb.sheetnames)
                raise ValueError(
                    f"Sheet '{sheet_name}' not found in {path.name}. Available: {names}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
        return [list(r) if r is not None else [] for r in ws.values]
    finally:
        wb.close()


def _read_table(path: Path, sheet_name: str = "") -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel(path, sheet_name=sheet_name)
    return _read_csv(path)


def _pick_value(
    raw_row: dict[str, str],
    normalized_alias_map: dict[str, list[str]],
    target: str,
    normalized_lookup: dict[str, str],
) -> str:
    aliases = normalized_alias_map.get(target, [])
    for alias in aliases:
        if alias in normalized_lookup:
            original = normalized_lookup[alias]
            return _clean_text(raw_row.get(original, ""))
    return ""


def _normalize_mapping(
    field_map: dict[str, list[str]],
) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for target, aliases in field_map.items():
        normalized_aliases = [_norm_header(a) for a in aliases]
        out[target] = normalized_aliases
    return out


def _pick_numeric_from_aliases(
    raw_row: dict[str, str],
    normalized_lookup: dict[str, str],
    aliases: list[str],
    min_val: float,
    max_val: float,
) -> float | None:
    for alias in aliases:
        key = _norm_header(alias)
        if key not in normalized_lookup:
            continue
        raw_val = raw_row.get(normalized_lookup[key], "")
        v = _to_float(raw_val)
        if v is None:
            continue
        if min_val <= v <= max_val:
            return v
    return None


def _repair_row_for_p3(
    out: dict[str, str],
    raw_row: dict[str, str],
    normalized_lookup: dict[str, str],
) -> None:
    # 清理非数据行（标题/注释行）
    rock_name = _clean_text(out.get("rock_type", "")).lower()
    non_data_tokens = {
        "rock type",
        "appendix a seismic velocity data",
        "numbers in yellow are expolated values",
    }
    if rock_name in non_data_tokens:
        out["rock_type"] = ""

    if _to_float(out.get("density", "")) is None:
        d = _pick_numeric_from_aliases(
            raw_row,
            normalized_lookup,
            aliases=["density_g_cm_3", "density_g_cm3", "felsic_or_mafic", "rock_facies"],
            min_val=1.5,
            max_val=4.0,
        )
        if d is not None:
            out["density"] = f"{d:.3f}"
            out["density_unit"] = out.get("density_unit") or "g/cm3"

    if _to_float(out.get("vp", "")) is None:
        vp = _pick_numeric_from_aliases(
            raw_row,
            normalized_lookup,
            aliases=["vp_km_s", "vp", "sio2_wt", "density_g_cm_3", "felsic_or_mafic"],
            min_val=4.0,
            max_val=9.0,
        )
        if vp is not None:
            out["vp"] = f"{vp:.3f}"

    if _to_float(out.get("vs", "")) is None:
        vs = _pick_numeric_from_aliases(
            raw_row,
            normalized_lookup,
            aliases=["vs", "vs_data", "rock_facies", "sio2_wt"],
            min_val=2.0,
            max_val=5.5,
        )
        if vs is not None:
            out["vs"] = f"{vs:.3f}"


def _to_template_row(**kwargs: str) -> dict[str, str]:
    row = {col: "" for col in TEMPLATE_COLUMNS}
    for k, v in kwargs.items():
        if k in row:
            row[k] = _clean_text(v)
    return row


def _find_best_col(header: list[Any], candidates: list[str]) -> int:
    best_idx = -1
    best_score = -1
    norm_targets = [_norm_header(c) for c in candidates]
    for idx, cell in enumerate(header):
        key = _norm_header(cell)
        if not key:
            continue
        score = 0
        for target in norm_targets:
            if target and target in key:
                score += len(target)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _safe_cell(row: list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return ""
    return row[idx]


def _looks_like_rock_name(name: str) -> bool:
    s = _clean_text(name).lower()
    if not s:
        return False
    bad_tokens = [
        "appendix",
        "numbers in yellow",
        "sample discription",
        "pressure",
        "data master",
        "these ",
    ]
    return not any(t in s for t in bad_tokens)


def _extract_reference_text(row: list[Any]) -> str:
    for cell in reversed(row):
        text = _clean_text(cell)
        if not text:
            continue
        # Support references like "Christensen 1966a"
        if re.search(r"\b(?:19|20)\d{2}[a-zA-Z]?\b", text):
            return text
    return ""


def _normalize_rock_name_518(name: str) -> str:
    s = _clean_text(name)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    # 统一大小写，减少 eclogite / Eclogite 等重复类别
    return " ".join(part.capitalize() for part in s.split(" "))


def _collect_pressure_columns(
    p_row: list[Any],
    start_idx: int,
    end_idx: int,
) -> list[tuple[int, float]]:
    out: list[tuple[int, float]] = []
    for c in range(max(0, start_idx), max(0, end_idx)):
        p = _to_float(_safe_cell(p_row, c))
        if p is None:
            continue
        if 0.0 < p <= 5.0:
            out.append((c, p))
    return out


def _pressure_in_window(pressure_gpa: float, min_gpa: float | None, max_gpa: float | None) -> bool:
    if min_gpa is not None and pressure_gpa < min_gpa:
        return False
    if max_gpa is not None and pressure_gpa > max_gpa:
        return False
    return True


def _transform_p3_518(
    source_path: Path,
    output_csv: Path,
    sheet_name: str,
    min_pressure_gpa: float | None = None,
    max_pressure_gpa: float | None = None,
) -> tuple[int, int]:
    all_rows = _read_excel_all_rows(source_path, sheet_name=sheet_name or "Data Master")
    if not all_rows:
        with output_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=TEMPLATE_COLUMNS)
            writer.writeheader()
        return 0, 0

    # 定位主表头行
    header_idx = 0
    best = -1
    for i, row in enumerate(all_rows[:120]):
        joined = " ".join(_clean_text(c).lower() for c in row if _clean_text(c))
        score = 0
        if "rock type" in joined:
            score += 3
        if "density" in joined:
            score += 2
        if "vp" in joined:
            score += 2
        if "sample" in joined:
            score += 1
        if score > best:
            best = score
            header_idx = i

    header_row = all_rows[header_idx]
    rock_col = _find_best_col(header_row, ["rock type"])
    density_col = _find_best_col(header_row, ["density"])
    felsic_mafic_col = _find_best_col(header_row, ["felsic or mafic", "felsic_mafic"])
    facies_col = _find_best_col(header_row, ["rock facies", "facies"])
    sio2_col = _find_best_col(header_row, ["sio2", "sio2 wt", "sio2, wt.%"])
    vp_anchor_col = _find_best_col(header_row, ["vp"])
    vs_anchor_col = _find_best_col(header_row, ["vs"])
    data_quality_col = _find_best_col(header_row, ["data quality"])

    # 压力行（GPa）与数值行
    pressure_label_idx = -1
    for i in range(header_idx, min(header_idx + 20, len(all_rows))):
        joined = " ".join(_clean_text(c).lower() for c in all_rows[i] if _clean_text(c))
        if "pressure" in joined and "gpa" in joined:
            pressure_label_idx = i
            break

    pressure_values_idx = pressure_label_idx + 1 if pressure_label_idx >= 0 else -1
    vp_pressure_cols: list[tuple[int, float]] = []
    vs_pressure_cols: list[tuple[int, float]] = []
    if pressure_values_idx >= 0 and pressure_values_idx < len(all_rows):
        p_row = all_rows[pressure_values_idx]
        vp_end = vs_anchor_col if vs_anchor_col > vp_anchor_col else len(p_row)
        vs_end = data_quality_col if data_quality_col > vs_anchor_col else len(p_row)
        vp_pressure_cols = _collect_pressure_columns(p_row, vp_anchor_col, vp_end)
        vs_pressure_cols = _collect_pressure_columns(p_row, vs_anchor_col, vs_end)

    if not vp_pressure_cols and vp_anchor_col >= 0:
        vp_pressure_cols = [(vp_anchor_col, 0.02)]
    vs_pressure_map: dict[float, int] = {p: c for c, p in vs_pressure_cols}

    data_start = max(header_idx + 1, pressure_values_idx + 1 if pressure_values_idx >= 0 else 0)
    out_rows: list[dict[str, str]] = []
    skipped_empty = 0
    current_source = "Huang2012_AppendixA"
    for row in all_rows[data_start:]:
        rock_type = _normalize_rock_name_518(_safe_cell(row, rock_col))
        if not _looks_like_rock_name(rock_type):
            skipped_empty += 1
            continue

        density = _to_float(_safe_cell(row, density_col))
        if density is None or not (1.5 <= density <= 4.0):
            candidates = [_to_float(c) for c in row]
            density = next((x for x in candidates if x is not None and 1.5 <= x <= 4.0), None)

        extracted_source = _extract_reference_text(row)
        if extracted_source:
            current_source = extracted_source
        source = current_source
        felsic_or_mafic = _clean_text(_safe_cell(row, felsic_mafic_col))
        rock_facies = _clean_text(_safe_cell(row, facies_col))
        sio2 = _to_float(_safe_cell(row, sio2_col))
        for col_idx, pressure_gpa in vp_pressure_cols:
            if not _pressure_in_window(pressure_gpa, min_pressure_gpa, max_pressure_gpa):
                continue
            vp = _to_float(_safe_cell(row, col_idx))
            if vp is None or not (4.0 <= vp <= 9.0):
                continue
            if density is None:
                continue

            vs_value = None
            vs_col_idx = vs_pressure_map.get(pressure_gpa)
            if vs_col_idx is not None:
                candidate_vs = _to_float(_safe_cell(row, vs_col_idx))
                if candidate_vs is not None and 2.0 <= candidate_vs <= 5.5:
                    vs_value = candidate_vs

            out_rows.append(
                _to_template_row(
                    rock_type=rock_type,
                    density=f"{density:.3f}",
                    density_unit="g/cm3",
                    vp=f"{vp:.3f}",
                    vs=f"{vs_value:.3f}" if vs_value is not None else "",
                    pressure=f"{pressure_gpa:.3f}",
                    pressure_unit="GPa",
                    temperature="25",
                    felsic_or_mafic=felsic_or_mafic,
                    rock_facies=rock_facies,
                    sio2_wt=f"{sio2:.3f}" if sio2 is not None else "",
                    source=source,
                    method="compiled_518_strict",
                )
            )

    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TEMPLATE_COLUMNS)
        writer.writeheader()
        writer.writerows(out_rows)
    return len(out_rows), skipped_empty


def transform(
    source_csv: Path,
    output_csv: Path,
    mapping_json: Path,
    preset: str,
    sheet_name: str = "",
    min_pressure_gpa: float | None = None,
    max_pressure_gpa: float | None = None,
) -> tuple[int, int]:
    if preset == "p3_518":
        return _transform_p3_518(
            source_csv,
            output_csv,
            sheet_name=sheet_name,
            min_pressure_gpa=min_pressure_gpa,
            max_pressure_gpa=max_pressure_gpa,
        )

    cfg = _load_json(mapping_json)
    presets = cfg.get("presets", {})
    if preset not in presets:
        available = ", ".join(sorted([*presets, "p3_518"]))
        raise ValueError(f"Unknown preset '{preset}'. Available: {available}")

    preset_cfg = presets[preset]
    field_map = _normalize_mapping(preset_cfg.get("field_map", {}))
    defaults = {
        k: _clean_text(v) for k, v in (preset_cfg.get("defaults", {}) or {}).items()
    }

    headers, rows = _read_table(source_csv, sheet_name=sheet_name)
    normalized_lookup = {_norm_header(h): h for h in headers}

    transformed_rows: list[dict[str, str]] = []
    skipped_empty = 0
    for raw in rows:
        out = {col: "" for col in TEMPLATE_COLUMNS}
        for col in TEMPLATE_COLUMNS:
            out[col] = _pick_value(raw, field_map, col, normalized_lookup)
            if not out[col] and col in defaults:
                out[col] = defaults[col]

        if preset == "p3":
            _repair_row_for_p3(out, raw, normalized_lookup)

        # Skip fully empty rows
        if not any(_clean_text(v) for v in out.values()):
            skipped_empty += 1
            continue
        transformed_rows.append(out)

    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TEMPLATE_COLUMNS)
        writer.writeheader()
        writer.writerows(transformed_rows)

    return len(transformed_rows), skipped_empty


def _report_sources_from_output(output_csv: Path, top_n: int = 20) -> None:
    counter: Counter[str] = Counter()
    with output_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            source = _clean_text(row.get("source", ""))
            if source:
                counter[source] += 1

    print(f"source_unique={len(counter)}")
    for source, count in counter.most_common(max(1, top_n)):
        print(f"source_count[{source}]={count}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Transform external source CSV into external_rock_data_template.csv format."
        )
    )
    parser.add_argument("--source-csv", required=True, help="Path to source CSV.")
    parser.add_argument("--preset", required=True, help="Mapping preset name.")
    parser.add_argument(
        "--mapping-json",
        default=str(Path(__file__).with_name("rock_source_field_mappings.json")),
        help="Mapping config JSON path.",
    )
    parser.add_argument(
        "--out",
        default=str(
            Path(__file__).parent / "rockdata" / "external_rock_data_template.csv"
        ),
        help="Output CSV path (template format).",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Excel sheet name for .xlsx/.xlsm input (optional).",
    )
    parser.add_argument(
        "--min-pressure-gpa",
        type=float,
        default=None,
        help="Only for p3_518: keep samples with pressure >= this value (GPa).",
    )
    parser.add_argument(
        "--max-pressure-gpa",
        type=float,
        default=None,
        help="Only for p3_518: keep samples with pressure <= this value (GPa).",
    )
    parser.add_argument(
        "--report-sources",
        action="store_true",
        help="Print source distribution summary after conversion.",
    )
    parser.add_argument(
        "--report-top",
        type=int,
        default=20,
        help="Top N sources to print when --report-sources is enabled.",
    )
    args = parser.parse_args()

    source_csv = Path(args.source_csv).expanduser().resolve()
    mapping_json = Path(args.mapping_json).expanduser().resolve()
    out = Path(args.out).expanduser().resolve()

    if not source_csv.exists():
        raise FileNotFoundError(f"Source CSV not found: {source_csv}")
    if not mapping_json.exists():
        raise FileNotFoundError(f"Mapping JSON not found: {mapping_json}")
    out.parent.mkdir(parents=True, exist_ok=True)

    converted, skipped_empty = transform(
        source_csv=source_csv,
        output_csv=out,
        mapping_json=mapping_json,
        preset=args.preset.strip(),
        sheet_name=args.sheet.strip(),
        min_pressure_gpa=args.min_pressure_gpa,
        max_pressure_gpa=args.max_pressure_gpa,
    )
    print(f"prepared_csv={out}")
    print(f"converted_rows={converted}")
    print(f"skipped_empty_rows={skipped_empty}")
    if args.report_sources:
        _report_sources_from_output(out, top_n=args.report_top)


if __name__ == "__main__":
    main()
